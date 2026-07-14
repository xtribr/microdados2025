#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gera TRI_COP30_BAM_ENEM2025.xlsx — todos os parâmetros TRI (a, b, c) oficiais
do INEP para os cadernos da aplicação COP30/BAM (Belém-Ananindeua-Marituba),
publicados em ITENS_PROVA_2025.csv sob os códigos 1499-1538 (família própria,
zero itens em comum com Regular e Reaplicação).

Fonte: Microdados ENEM 2025 / INEP (dados públicos). Nenhum valor estimado.
"""
import csv
import math
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

D = Path(__file__).resolve().parent
BASE = Path("/Volumes/Kingston 1/microdados_enem_2025")
ITENS = BASE / "DADOS/ITENS_PROVA_2025.csv"
PCT = BASE / "estudo_marista_belem/itens_cop30.csv"
OUT = D / "TRI_COP30_BAM_ENEM2025.xlsx"

AREAS = ["LC", "CH", "CN", "MT"]
NOME_AREA = {"LC": "Linguagens", "CH": "Ciências Humanas",
             "CN": "Ciências da Natureza", "MT": "Matemática"}
# cadernos principais (4 cores) da família COP30/BAM em ITENS_PROVA
MAIN = {"CH": ["1520", "1521", "1522", "1523"], "LC": ["1529", "1530", "1531", "1532"],
        "MT": ["1502", "1503", "1504", "1505"], "CN": ["1511", "1512", "1513", "1514"]}
FAMILIA_BAM_ITENS = [str(c) for c in range(1499, 1539)]
# mapa validado por gabarito: código no RESULTADOS (BAM2) -> código em ITENS
RES2ITEM = {
    "CH": {"1583": "1520", "1584": "1521", "1585": "1522", "1586": "1523",
           "1587": "1522", "1631": "1522"},
    "LC": {"1595": "1529", "1596": "1530", "1597": "1531", "1598": "1532",
           "1599": "1532", "1632": "1532"},
    "MT": {"1607": "1502", "1608": "1503", "1609": "1504", "1610": "1505",
           "1611": "1505", "1633": "1505"},
    "CN": {"1619": "1511", "1620": "1512", "1621": "1514", "1622": "1513",
           "1623": "1513", "1634": "1513"},
}
ROTULO_RES = {"1583": "Azul", "1584": "Amarela", "1585": "Branca", "1586": "Verde",
              "1587": "Laranja-Ampliada", "1631": "Laranja-At.Especializado",
              "1595": "Azul", "1596": "Amarela", "1597": "Verde", "1598": "Branca",
              "1599": "Laranja-Ampliada", "1632": "Laranja-At.Especializado",
              "1607": "Azul", "1608": "Amarela", "1609": "Verde", "1610": "Cinza",
              "1611": "Laranja-Ampliada", "1633": "Laranja-At.Especializado",
              "1619": "Azul", "1620": "Amarela", "1621": "Verde", "1622": "Cinza",
              "1623": "Laranja-Ampliada", "1634": "Laranja-At.Especializado"}

INK = Font(bold=True, color="FFFFFF")
HFILL = PatternFill("solid", fgColor="1F4E78")
CTR = Alignment("center", "center")
THIN = Border(*[Side(style="thin", color="D9D9D9")] * 4)


def pf(v):
    try:
        return float(v) if v not in (None, "") else None
    except ValueError:
        return None


def p_theta0(a, b, c):
    return c + (1 - c) / (1 + math.exp(1.7 * a * b))


def carregar():
    rows = []
    for r in csv.DictReader(open(ITENS, encoding="latin-1"), delimiter=";"):
        if r["CO_PROVA"].strip() in FAMILIA_BAM_ITENS:
            rows.append({
                "area": r["SG_AREA"].strip(), "co_prova": r["CO_PROVA"].strip(),
                "cor": r["TX_COR"].strip(), "pos": int(r["CO_POSICAO"]),
                "co_item": r["CO_ITEM"].strip(), "gab": (r["TX_GABARITO"] or "").strip(),
                "hab": r["CO_HABILIDADE"].strip(), "lingua": (r["TP_LINGUA"] or "").strip().split(".")[0],
                "A": pf(r["NU_PARAM_A"]), "B": pf(r["NU_PARAM_B"]), "C": pf(r["NU_PARAM_C"]),
                "adapt": r["IN_ITEM_ADAPTADO"].strip(), "aban": r["IN_ITEM_ABAN"].strip(),
            })
    return rows


def pct_coorte():
    p = {}
    for r in csv.DictReader(open(PCT, encoding="utf-8")):
        p[r["co_item"]] = (pf(r["pct_acerto_cop30"]), int(float(r["respostas_cop30"] or 0)))
    return p


def estilo_header(ws, ncols):
    for j in range(1, ncols + 1):
        c = ws.cell(1, j)
        c.font = INK
        c.fill = HFILL
        c.alignment = CTR
        c.border = THIN
    ws.freeze_panes = "A2"


def main():
    rows = carregar()
    pcts = pct_coorte()
    wb = openpyxl.Workbook()

    # ---------- LEIA-ME ----------
    ws = wb.active
    ws.title = "LEIA-ME"
    L = [
        ["PARÂMETROS TRI OFICIAIS (a, b, c) — ENEM 2025, APLICAÇÃO COP30/BAM (Belém, Ananindeua e Marituba)"],
        [""],
        ["O que é isto",
         "Os 3 parâmetros TRI de TODOS os itens dos cadernos aplicados na COP30/BAM, extraídos sem alteração do arquivo público ITENS_PROVA_2025.csv (Microdados ENEM 2025 / INEP)."],
        ["Por que a comunidade não encontrou",
         "No RESULTADOS_2025, os candidatos da COP30 têm códigos de prova 1583-1634 (rótulo \"BAM2\" no Dicionário). Esses códigos têm 0 linhas no ITENS_PROVA — quem procura por eles conclui que os parâmetros não existem. Mas o INEP publicou os MESMOS cadernos no ITENS_PROVA sob outra numeração: 1499-1538."],
        ["Prova de que 1499-1538 é a prova da COP30 (3 evidências independentes)",
         "1) O gabarito desses cadernos, posição por posição, é IDÊNTICO ao TX_GABARITO que o INEP grava no RESULTADOS para os ~66 mil candidatos BAM2 — zero divergências em todos os cadernos e áreas. 2) Zero CO_ITEM em comum com a prova Regular (1447-1486) e com a Reaplicação (1539-1582) — são três provas disjuntas. 3) A estrutura de cores bate (MT/CN: Azul/Amarela/Verde/Cinza; CH/LC: Azul/Amarela/Verde/Branca), além de variantes Laranja e Leitor de Tela."],
        ["Como verificar por conta própria",
         "Roda o script verificar_tri_cop30.R (ou o passo manual do COMO_VERIFICAR.md): basta o par de arquivos públicos RESULTADOS_2025.csv + ITENS_PROVA_2025.csv."],
        ["Abas",
         "\"Mapeamento códigos\" = ponte RESULTADOS↔ITENS por caderno. \"MT/CN/CH/LC\" = itens dos 4 cadernos principais por área, com a/b/c, habilidade, gabarito e dificuldade na escala ENEM (b×100+500). \"Itens únicos\" = os 185 itens objetivos sem repetição, com %acerto observado na coorte COP30 (~62-66 mil presentes)."],
        ["Observações",
         "A prova COP30/BAM não teve item anulado (IN_ITEM_ABAN=0 em todos). TP_LINGUA em LC: 0=Inglês, 1=Espanhol, vazio=tronco comum. Parâmetros na métrica publicada pelo INEP (3PL; dificuldade ENEM = b×100+500)."],
        ["Fonte",
         "Microdados ENEM 2025 e Dicionário / INEP (gov.br/inep). Compilação: Prof. Alexandre Emerson (XTRI). Nenhum valor foi estimado, ajustado ou imputado."],
    ]
    for row in L:
        ws.append(row)
    ws["A1"].font = Font(bold=True, size=13)
    for i in range(3, 10):
        ws.cell(i, 1).font = Font(bold=True)
        ws.cell(i, 2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 130

    # ---------- Mapeamento códigos ----------
    ws = wb.create_sheet("Mapeamento códigos")
    ws.append(["Área", "Caderno", "Código no RESULTADOS (aluno)", "Código no ITENS_PROVA (parâmetros)",
               "Validação do gabarito"])
    for a in ["CH", "LC", "MT", "CN"]:
        for rc, ic in sorted(RES2ITEM[a].items()):
            ws.append([NOME_AREA[a], ROTULO_RES[rc], int(rc), int(ic),
                       "idêntico posição a posição (0 divergências)"])
    estilo_header(ws, 5)
    for i in range(2, ws.max_row + 1):
        for j in range(1, 6):
            ws.cell(i, j).border = THIN
            if j in (3, 4):
                ws.cell(i, j).alignment = CTR
    for j, w in enumerate([22, 26, 28, 32, 40], 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # ---------- abas por área (4 cadernos principais) ----------
    for a in ["MT", "CN", "CH", "LC"]:
        ws = wb.create_sheet(a)
        head = ["CO_PROVA (ITENS)", "Cor", "CO_PROVA (RESULTADOS)", "Posição", "CO_ITEM",
                "Gabarito", "Habilidade", "Língua", "Param A", "Param B", "Param C",
                "Dificuldade ENEM (b×100+500)"]
        ws.append(head)
        inv = {ic: rc for rc, ic in RES2ITEM[a].items() if ROTULO_RES[rc] in
               ("Azul", "Amarela", "Verde", "Cinza", "Branca")}
        for r in sorted((x for x in rows if x["area"] == a and x["co_prova"] in MAIN[a]),
                        key=lambda x: (x["co_prova"], x["pos"], x["lingua"])):
            lingua = {"0": "Inglês", "1": "Espanhol"}.get(r["lingua"], "")
            ws.append([int(r["co_prova"]), r["cor"], int(inv.get(r["co_prova"], 0)) or "",
                       r["pos"], int(r["co_item"]), r["gab"], f"H{int(r['hab'])}", lingua,
                       r["A"], r["B"], r["C"],
                       round(r["B"] * 100 + 500, 1) if r["B"] is not None else ""])
        estilo_header(ws, len(head))
        ws.auto_filter.ref = f"A1:{get_column_letter(len(head))}{ws.max_row}"
        for j, w in enumerate([18, 12, 22, 10, 12, 10, 12, 11, 11, 11, 11, 26], 1):
            ws.column_dimensions[get_column_letter(j)].width = w
        for i in range(2, ws.max_row + 1):
            for j in range(1, len(head) + 1):
                ws.cell(i, j).border = THIN
                if j not in (2,):
                    ws.cell(i, j).alignment = CTR

    # ---------- Itens únicos ----------
    ws = wb.create_sheet("Itens únicos")
    head = ["Área", "CO_ITEM", "Habilidade", "Língua", "Gabarito", "Param A", "Param B",
            "Param C", "Dificuldade ENEM", "P(acerto) modelo em θ=0",
            "% acerto observado (coorte COP30)", "N respostas (coorte)", "Observação"]
    ws.append(head)
    vistos = set()
    uniq = []
    for r in rows:
        if r["co_prova"] in MAIN[r["area"]] and r["co_item"] not in vistos:
            vistos.add(r["co_item"])
            uniq.append(dict(r, obs=""))
    # item substituto dos cadernos adaptados (existe fora dos 4 cadernos principais)
    for r in rows:
        if r["co_item"] not in vistos:
            vistos.add(r["co_item"])
            uniq.append(dict(r, obs="só nos cadernos adaptados (Laranja/Leitor de Tela) — "
                                     "substitui o item da mesma posição"))
    uniq.sort(key=lambda r: (AREAS.index(r["area"]), r["pos"], r["lingua"]))
    for r in uniq:
        pobs, n = pcts.get(r["co_item"], (None, None))
        pm = p_theta0(r["A"], r["B"], r["C"]) if r["A"] is not None else None
        lingua = {"0": "Inglês", "1": "Espanhol"}.get(r["lingua"], "")
        ws.append([NOME_AREA[r["area"]], int(r["co_item"]), f"H{int(r['hab'])}", lingua,
                   r["gab"], r["A"], r["B"], r["C"],
                   round(r["B"] * 100 + 500, 1) if r["B"] is not None else "",
                   round(pm, 4) if pm is not None else "", pobs, n, r["obs"]])
    estilo_header(ws, len(head))
    ws.auto_filter.ref = f"A1:{get_column_letter(len(head))}{ws.max_row}"
    for j, w in enumerate([22, 12, 12, 11, 10, 11, 11, 11, 18, 22, 30, 20, 52], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    for i in range(2, ws.max_row + 1):
        for j in range(1, len(head) + 1):
            ws.cell(i, j).border = THIN
            if 1 < j < len(head):
                ws.cell(i, j).alignment = CTR

    # ---------- Inventário da família (40 cadernos) ----------
    ws = wb.create_sheet("Inventário 1499-1538")
    ws.append(["CO_PROVA (ITENS)", "Área", "Cor", "Nº itens", "Itens adaptados",
               "Com parâmetros a-b-c"])
    inv = defaultdict(lambda: {"n": 0, "ad": 0, "pa": 0, "area": "", "cor": ""})
    for r in rows:
        d = inv[r["co_prova"]]
        d["n"] += 1
        d["ad"] += (r["adapt"] == "1")
        d["pa"] += (r["A"] is not None)
        d["area"], d["cor"] = r["area"], d["cor"] or r["cor"]
    for cp in sorted(inv, key=int):
        d = inv[cp]
        ws.append([int(cp), NOME_AREA[d["area"]], d["cor"], d["n"], d["ad"],
                   f"{d['pa']}/{d['n']}"])
    estilo_header(ws, 6)
    for i in range(2, ws.max_row + 1):
        for j in range(1, 7):
            ws.cell(i, j).border = THIN
            ws.cell(i, j).alignment = CTR
    for j, w in enumerate([20, 24, 16, 10, 16, 22], 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    wb.save(OUT)
    n_area = {a: sum(1 for r in rows if r["area"] == a and r["co_prova"] in MAIN[a]) for a in AREAS}
    print(f"OK: {OUT.name}")
    print(f"linhas por área (4 cadernos): {n_area} | itens únicos: {len(uniq)}")


if __name__ == "__main__":
    main()
