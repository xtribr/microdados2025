#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Exporta distratores_itens.csv para um .xlsx formatado (padrão XTRI) para uso do professor."""
import csv
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

D = Path(__file__).resolve().parent
rows = list(csv.DictReader((D / "distratores_itens.csv").open(encoding="utf-8")))

NOMES_AREA = {"LC": "Linguagens", "CH": "Ciências Humanas", "CN": "Ciências da Natureza", "MT": "Matemática"}
FILLC = {"LC": "E8F0FE", "CH": "FCE8E6", "CN": "E6F4EA", "MT": "FEF7E0"}

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Distratores"
H = ["Área", "Posição", "Nº (dia)", "Habilidade", "Descrição da habilidade", "Gabarito", "% Acerto",
     "Distrator campeão", "% entre os que erraram", "% da base (todos)", "Lift vs. 25% (uniforme)",
     "Dificuldade TRI"]
ws.append(H)
hf = PatternFill("solid", fgColor="1F4E78")
hfont = Font(bold=True, color="FFFFFF")
ctr = Alignment("center", "center")
lft = Alignment("left", "center", wrap_text=True)
b = Border(*[Side(style="thin", color="D9D9D9")] * 4)
for i, _ in enumerate(H, 1):
    c = ws.cell(1, i)
    c.fill = hf
    c.font = hfont
    c.alignment = ctr
    c.border = b

rows.sort(key=lambda r: -float(r["lift_vs_uniforme"]))
for r in rows:
    area = r["area"]
    ws.append([
        NOMES_AREA[area], int(r["posicao"]), int(r["co_posicao_abs"]),
        r["habilidade_cod"], r["habilidade_desc"], r["gabarito"],
        float(r["pct_acerto"]) / 100, r["distrator_campeao"],
        float(r["pct_campeao_entre_errados"]) / 100, float(r["pct_campeao_da_base"]) / 100,
        float(r["lift_vs_uniforme"]), r["dif_tri"],
    ])
    ri = ws.max_row
    for ci in range(1, len(H) + 1):
        cell = ws.cell(ri, ci)
        cell.border = b
        cell.alignment = ctr if ci in (1, 2, 3, 6, 8, 12) else lft
        if ci == 1:
            cell.fill = PatternFill("solid", fgColor=FILLC.get(area, "FFFFFF"))
    ws.cell(ri, 7).number_format = "0.0%"
    ws.cell(ri, 9).number_format = "0.0%"
    ws.cell(ri, 10).number_format = "0.0%"
    ws.cell(ri, 11).number_format = "0.00\"x\""
    if float(r["pct_campeao_da_base"]) > float(r["pct_acerto"]):
        ws.cell(ri, 8).font = Font(bold=True, color="C00000")

widths = [16, 9, 9, 12, 46, 9, 10, 16, 16, 14, 14, 12]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(H))}{ws.max_row}"

# aba resumo
s = wb.create_sheet("Resumo")
s.append(["MAPA DE DISTRATORES — ENEM 2025"])
s["A1"].font = Font(bold=True, size=13)
s.append([])
s.append(["Total de itens analisados", len(rows)])
s.append(["Lift médio (todas as áreas)", round(sum(float(r["lift_vs_uniforme"]) for r in rows) / len(rows), 2)])
s.append(["Itens com pegadinha clara (lift ≥ 1,5x)", sum(1 for r in rows if float(r["lift_vs_uniforme"]) >= 1.5)])
s.append(["Itens onde o distrator supera o gabarito em popularidade",
           sum(1 for r in rows if float(r["pct_campeao_da_base"]) > float(r["pct_acerto"]))])
s.column_dimensions["A"].width = 50
wb.save(D / "MAPA_DISTRATORES_ENEM2025.xlsx")
print("ok: MAPA_DISTRATORES_ENEM2025.xlsx")
