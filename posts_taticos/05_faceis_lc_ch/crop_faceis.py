#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Recorta o print de cada questão FÁCIL de LC+CH (dif TRI <= 548) do caderno Azul dia 1.
Usa pymupdf: acha o header 'QUESTÃO N', detecta a coluna (prova em 2 colunas) e recorta
até o próximo header na mesma coluna. Saída: thumbnails PNG + faceis_meta.json."""
import json
import re
from pathlib import Path
import fitz
import openpyxl

BASE = Path("/Volumes/Kingston 1/microdados_enem_2025")
PDF = BASE / "PROVAS E GABARITOS/ENEM_2025_P1_CAD_01_DIA_1_AZUL.pdf"
OUT = BASE / "posts_taticos/05_faceis_lc_ch"
CROPS = OUT / "crops"; CROPS.mkdir(exist_ok=True)
DIF_MAX = 548          # faixa "Fácil" (b*100+500 <= 548)
DPI = 200

# ---- easy items LC+CH (não anulados, não-língua p/ recorte limpo) ----
wb = openpyxl.load_workbook(BASE / "TRI_ITENS_AZUL_ENEM2025.xlsx", read_only=True)
ws = wb.active
hdr = list(next(ws.iter_rows(values_only=True)))
sigla = {"Linguagens": "LC", "Ciências Humanas": "CH"}
itens = []
for r in ws.iter_rows(min_row=2, values_only=True):
    d = dict(zip(hdr, r))
    if d["Área"] not in sigla or d.get("Anulada"):
        continue
    if (d["Língua"] or ""):        # pula itens de língua estrangeira (header duplicado no caderno)
        continue
    dif = d["TRI dificuldade (b×100+500)"]
    if dif > DIF_MAX:
        continue
    itens.append({"area": sigla[d["Área"]], "n": int(d["Nº"]), "hab": f"{sigla[d['Área']]}H{d['Hab.']}",
                  "pct": round(d["% acerto observado"], 1), "dif": round(dif), "gab": d["Gab."]})
itens.sort(key=lambda x: (x["area"] != "LC", -x["pct"]))
print(f"fáceis (dif<= {DIF_MAX}, sem língua): {sum(1 for i in itens if i['area']=='LC')} LC + "
      f"{sum(1 for i in itens if i['area']=='CH')} CH")

# ---- mapear questão -> página ----
doc = fitz.open(PDF)
qpage = {}
for pi in range(doc.page_count):
    for m in re.finditer(r"QUEST[ÃA]O\s*0*(\d{1,3})", doc[pi].get_text().upper()):
        qpage.setdefault(int(m.group(1)), pi)


def crop_q(n):
    pi = qpage.get(n)
    if pi is None:
        return None
    page = doc[pi]
    Wp = page.rect.width
    words = page.get_text("words")  # x0,y0,x1,y1,word,block,line,wno
    # acha header "QUESTÃO" cujo próximo token é o número n
    heads = []
    for i, w in enumerate(words):
        if re.match(r"QUEST[ÃA]O", w[4].upper()):
            for j in range(i + 1, min(i + 3, len(words))):
                mm = re.match(r"0*(\d{1,3})$", words[j][4])
                if mm:
                    heads.append((int(mm.group(1)), w[0], w[1]))
                    break
    tgt = [h for h in heads if h[0] == n]
    if not tgt:
        return None
    _, hx, hy = tgt[0]
    left = hx < Wp / 2
    x0, x1 = (0.045 * Wp, 0.492 * Wp) if left else (0.508 * Wp, 0.915 * Wp)
    ybot = page.rect.height * 0.94
    for (hn, x, y) in heads:
        if (x < Wp / 2) == left and y > hy + 6 and y < ybot:
            ybot = y
    ytop = max(0, hy - 6)
    mat = fitz.Matrix(DPI / 72.0, DPI / 72.0)
    clip = fitz.Rect(x0, ytop, x1, ybot)
    pix = page.get_pixmap(matrix=mat, clip=clip)
    fp = CROPS / f"q{n:02d}_{'LC' if n <= 45 else 'CH'}.png"
    pix.save(str(fp))
    return str(fp), (pix.width, pix.height)


ok = 0
for it in itens:
    res = crop_q(it["n"])
    if res:
        it["crop"] = res[0]; it["size"] = res[1]; ok += 1
        print(f"  Q{it['n']:02d} {it['area']} pg{qpage[it['n']]+1} pct{it['pct']} dif{it['dif']} -> {Path(res[0]).name} {res[1]}")
    else:
        it["crop"] = None
        print(f"  !! Q{it['n']} ({it['area']}) não recortada")
(OUT / "faceis_meta.json").write_text(json.dumps({"dif_max": DIF_MAX, "itens": itens}, ensure_ascii=False, indent=2))
print(f"\n{ok}/{len(itens)} recortadas -> faceis_meta.json")
