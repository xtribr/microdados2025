#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Exporta TODOS os registros de PE (SG_UF_PROVA) do RESULTADOS_2025.csv para Excel:
aba LEIA-ME + aba DADOS (272.279 linhas × 70 colunas). Códigos como texto, notas como número."""
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.cell import WriteOnlyCell

SRC = "DADOS/RESULTADOS_2025.csv"
OUT = "MICRODADOS_ENEM2025_PE.xlsx"

wb = Workbook(write_only=True)

# ---- aba LEIA-ME ----
ws0 = wb.create_sheet("LEIA-ME")
info = [
    ["MICRODADOS ENEM 2025 — RECORTE PERNAMBUCO (PE)"],
    [""],
    ["O que é", "Todos os registros do arquivo oficial RESULTADOS_2025.csv com SG_UF_PROVA = PE (candidatos que fizeram a prova em Pernambuco)."],
    ["Linhas de dados", "272.279 candidatos"],
    ["Colunas", "70 (todas as colunas originais do INEP, sem alteração)"],
    ["Fonte", "Microdados ENEM 2025 / INEP — gov.br/inep > Dados Abertos > Microdados"],
    ["Recorte feito por", "Estudo XTRI — Prof. Alexandre Emerson (app.rankingenem.com)"],
    [""],
    ["COMO USAR"],
    ["1.", "A aba DADOS tem filtro automático no cabeçalho — use as setinhas para filtrar município, escola, presença etc."],
    ["2.", "Notas (NU_NOTA_...) já estão como número: dá para calcular média, máximo etc. direto."],
    ["3.", "Códigos (CO_ESCOLA, CO_MUNICIPIO...) estão como TEXTO de propósito — não converta, o Excel corrompe códigos longos."],
    ["4.", "O significado de cada coluna está no Dicionário oficial que acompanha os microdados do INEP."],
    [""],
    ["AVISOS IMPORTANTES"],
    ["•", "Nota 0 com presença 1 normalmente é redação zerada ou eliminação — confira TP_STATUS_REDACAO."],
    ["•", "TP_PRESENCA: 0 = faltou, 1 = presente, 2 = eliminado."],
    ["•", "TX_RESPOSTAS/TX_GABARITO são os vetores de resposta item a item (avançado — dá para corrigir a prova de cada aluno)."],
    ["•", "Alunos são anônimos: os microdados não têm nome, CPF nem data de nascimento completa."],
]
b = Font(bold=True)
for i, row in enumerate(info):
    cells = []
    for j, v in enumerate(row):
        c = WriteOnlyCell(ws0, value=v)
        if i == 0 or row[0] in ("COMO USAR", "AVISOS IMPORTANTES") or (j == 0 and len(row) > 1):
            c.font = b
        cells.append(c)
    ws0.append(cells)

# ---- aba DADOS ----
ws = wb.create_sheet("DADOS")
hdr_fill = PatternFill("solid", fgColor="1F4E78")
hdr_font = Font(bold=True, color="FFFFFF")

with open(SRC, encoding="latin-1", newline="") as f:
    reader = csv.reader(f, delimiter=";")
    head = next(reader)
    iuf = head.index("SG_UF_PROVA")
    num_idx = {i for i, h in enumerate(head) if h.startswith("NU_NOTA")}
    hcells = []
    for h in head:
        c = WriteOnlyCell(ws, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        hcells.append(c)
    ws.append(hcells)
    n = 0
    for row in reader:
        if row[iuf] != "PE":
            continue
        out = []
        for i, v in enumerate(row):
            if i in num_idx and v:
                try:
                    out.append(float(v.replace(",", ".")))
                except ValueError:
                    out.append(v)
            else:
                out.append(v if v != "" else None)
        ws.append(out)
        n += 1
        if n % 50_000 == 0:
            print(f"  {n:,} linhas...", flush=True)

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{'BR'}{n+1}"   # 70 colunas = BR
wb.save(OUT)
print(f"ok: {OUT} ({n:,} linhas de dados + cabeçalho)")
